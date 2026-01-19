from flask import Flask, request, jsonify, render_template
from flask_sqlalchemy import SQLAlchemy
from flask_jwt_extended import JWTManager, jwt_required, create_access_token, get_jwt_identity
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, date
from flask_cors import CORS
from sqlalchemy import Numeric, Text
import os
from dotenv import load_dotenv
from urllib.parse import quote_plus
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from flask import send_file
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

app = Flask(__name__)

# Enable CORS for React frontend
CORS(app, origins=['http://localhost:3000'])

# Load variables from .env if present
load_dotenv()

# Database URI builder with env-based configuration and safe fallback

def build_db_uri():
    # Highest priority: full DATABASE_URL (e.g., mysql+pymysql://user:pass@host:3306/db)
    db_url = os.getenv('DATABASE_URL')
    if db_url:
        return db_url

    # Next: compose from MYSQL_* variables
    mysql_user = os.getenv('MYSQL_USER')
    mysql_password = os.getenv('MYSQL_PASSWORD')
    mysql_host = os.getenv('MYSQL_HOST', 'localhost')
    mysql_port = os.getenv('MYSQL_PORT', '3306')
    mysql_db = os.getenv('MYSQL_DB')

    if mysql_user and mysql_password and mysql_db:
        return f"mysql+pymysql://{mysql_user}:{quote_plus(mysql_password)}@{mysql_host}:{mysql_port}/{mysql_db}"

    # Fallback: local SQLite DB placed next to this file
    project_dir = os.path.dirname(os.path.abspath(__file__))
    return os.getenv('SQLITE_URL', f"sqlite:///{os.path.join(project_dir, 'attendance.db')}")

# Configuration
app.config['SECRET_KEY'] = 'your-secret-key-change-this-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = build_db_uri()
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['JWT_SECRET_KEY'] = 'jwt-secret-string-change-this-in-production'
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=24)
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
}

# Small hint in logs about which backend is in use (no secrets printed)
try:
    print(f"Using database backend: {app.config['SQLALCHEMY_DATABASE_URI'].split('://')[0]}")
except Exception:
    pass

# Initialize extensions
db = SQLAlchemy(app)
jwt = JWTManager(app)

# Helper Functions
def log_audit_action(user_id, action, table_name=None, record_id=None, old_values=None, new_values=None, description=None):
    """Log audit action to database"""
    try:
        # Get IP address and user agent from request
        ip_address = request.remote_addr if request else None
        user_agent = request.headers.get('User-Agent') if request else None
        
        audit_log = AuditLog(
            user_id=user_id,
            action=action,
            table_name=table_name,
            record_id=record_id,
            old_values=old_values,
            new_values=new_values,
            ip_address=ip_address,
            user_agent=user_agent,
            description=description
        )
        db.session.add(audit_log)
        db.session.commit()
    except Exception as e:
        print(f"Audit log error: {e}")
        db.session.rollback()

def save_file_to_db(file_data, filename, file_type, description=None, related_table=None, related_id=None):
    """Save file data to database"""
    try:
        import os
        from werkzeug.utils import secure_filename
        
        # Generate secure filename
        secure_name = secure_filename(filename)
        file_size = len(file_data)
        
        # Determine MIME type based on extension
        mime_types = {
            '.pdf': 'application/pdf',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.xls': 'application/vnd.ms-excel',
            '.csv': 'text/csv',
            '.png': 'image/png',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg'
        }
        
        file_ext = os.path.splitext(filename)[1].lower()
        mime_type = mime_types.get(file_ext, 'application/octet-stream')
        
        file_storage = FileStorage(
            filename=secure_name,
            original_filename=filename,
            file_type=file_type,
            file_size=file_size,
            mime_type=mime_type,
            file_path=f'/uploads/{secure_name}',
            file_data=file_data,
            description=description,
            related_table=related_table,
            related_id=related_id,
            uploaded_by=get_jwt_identity()
        )
        
        db.session.add(file_storage)
        db.session.commit()
        return file_storage.id
        
    except Exception as e:
        print(f"File save error: {e}")
        db.session.rollback()
        return None

# JWT identity loader
@jwt.user_identity_loader
def user_identity_lookup(admin_id):
    return admin_id

@jwt.user_lookup_loader
def user_lookup_callback(_jwt_header, jwt_data):
    identity = jwt_data["sub"]
    return Admin.query.filter_by(id=int(identity)).one_or_none()

# Database Models
class Admin(db.Model):
    __tablename__ = 'admin'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=True)
    full_name = db.Column(db.String(100), nullable=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    last_login = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

class Department(db.Model):
    __tablename__ = 'departments'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    description = db.Column(Text, nullable=True)
    manager_id = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    
    employees = db.relationship('Employee', backref='department', lazy=True, foreign_keys='Employee.department_id')
    manager = db.relationship('Employee', foreign_keys=[manager_id], post_update=True)

class Employee(db.Model):
    __tablename__ = 'employees'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    phone = db.Column(db.String(20), nullable=True)
    address = db.Column(Text, nullable=True)
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'), nullable=True)
    position = db.Column(db.String(100), nullable=True)
    hire_date = db.Column(db.Date, nullable=True)
    salary = db.Column(Numeric(10, 2), nullable=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

class Attendance(db.Model):
    __tablename__ = 'attendance'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(20), nullable=False)  # 'present', 'absent', 'half_day', 'leave', 'overtime'
    check_in_time = db.Column(db.Time, nullable=True)
    check_out_time = db.Column(db.Time, nullable=True)
    total_hours = db.Column(Numeric(4, 2), nullable=True)
    overtime_hours = db.Column(Numeric(4, 2), default=0, nullable=True)
    notes = db.Column(Text, nullable=True)
    marked_by = db.Column(db.Integer, db.ForeignKey('admin.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    
    employee = db.relationship('Employee', backref=db.backref('attendance_records', lazy=True))
    admin = db.relationship('Admin', backref=db.backref('marked_attendance', lazy=True))
    
    __table_args__ = (db.UniqueConstraint('employee_id', 'date', name='unique_employee_date'),)

class Leave(db.Model):
    __tablename__ = 'leaves'
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=False)
    leave_type = db.Column(db.String(50), nullable=False)  # 'sick', 'vacation', 'personal', 'emergency'
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    days_count = db.Column(db.Integer, nullable=False)
    reason = db.Column(Text, nullable=True)
    status = db.Column(db.String(20), default='pending', nullable=False)  # 'pending', 'approved', 'rejected'
    approved_by = db.Column(db.Integer, db.ForeignKey('admin.id'), nullable=True)
    approved_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    
    employee = db.relationship('Employee', backref=db.backref('leave_requests', lazy=True))
    approver = db.relationship('Admin', backref=db.backref('approved_leaves', lazy=True))

class FileStorage(db.Model):
    __tablename__ = 'file_storage'
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    file_type = db.Column(db.String(50), nullable=False)  # 'excel', 'pdf', 'image', 'document'
    file_size = db.Column(db.Integer, nullable=False)
    mime_type = db.Column(db.String(100), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    file_data = db.Column(db.LargeBinary, nullable=True)  # Store file content in database
    description = db.Column(Text, nullable=True)
    related_table = db.Column(db.String(50), nullable=True)  # 'attendance', 'employee', 'report'
    related_id = db.Column(db.Integer, nullable=True)
    uploaded_by = db.Column(db.Integer, db.ForeignKey('admin.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    
    uploader = db.relationship('Admin', backref=db.backref('uploaded_files', lazy=True))

class Holiday(db.Model):
    __tablename__ = 'holidays'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    date = db.Column(db.Date, nullable=False)
    description = db.Column(Text, nullable=True)
    is_recurring = db.Column(db.Boolean, default=False, nullable=False)
    created_by = db.Column(db.Integer, db.ForeignKey('admin.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    
    creator = db.relationship('Admin', backref=db.backref('created_holidays', lazy=True))
    
    __table_args__ = (db.UniqueConstraint('name', 'date', name='unique_holiday_name_date'),)

class AuditLog(db.Model):
    __tablename__ = 'audit_logs'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('admin.id'), nullable=False)
    action = db.Column(db.String(100), nullable=False)  # 'CREATE', 'UPDATE', 'DELETE', 'LOGIN', 'EXPORT'
    table_name = db.Column(db.String(50), nullable=True)
    record_id = db.Column(db.Integer, nullable=True)
    old_values = db.Column(db.JSON, nullable=True)
    new_values = db.Column(db.JSON, nullable=True)
    ip_address = db.Column(db.String(45), nullable=True)
    user_agent = db.Column(db.String(255), nullable=True)
    description = db.Column(Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    
    user = db.relationship('Admin', backref=db.backref('audit_logs', lazy=True))

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/admin/login', methods=['POST'])
def admin_login():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'message': 'No data provided'}), 400
            
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return jsonify({'message': 'Username and password are required'}), 400
        
        admin = Admin.query.filter_by(username=username).first()
        
        if admin and admin.is_active and check_password_hash(admin.password_hash, password):
            # Update last login
            admin.last_login = datetime.utcnow()
            db.session.commit()
            
            access_token = create_access_token(identity=str(admin.id))
            
            # Log successful login
            log_audit_action(admin.id, 'LOGIN', None, None, None, 
                           {'login_time': datetime.utcnow().isoformat()}, 'Successful login')
            
            return jsonify({
                'access_token': access_token,
                'message': 'Login successful',
                'user': {
                    'id': admin.id,
                    'username': admin.username,
                    'full_name': admin.full_name,
                    'email': admin.email
                }
            }), 200
        else:
            # Log failed login attempt
            if admin:
                log_audit_action(admin.id, 'LOGIN_FAILED', None, None, None, 
                               {'username': username}, 'Failed login attempt')
            return jsonify({'message': 'Invalid credentials'}), 401
    except Exception as e:
        print(f"Login error: {e}")
        return jsonify({'message': 'Internal server error'}), 500

@app.route('/admin/test-token', methods=['GET'])
@jwt_required()
def verify_token_endpoint():
    """Test if JWT token is valid"""
    try:
        current_user_id = get_jwt_identity()
        admin = Admin.query.filter_by(id=int(current_user_id)).first()
        
        if admin and admin.is_active:
            return jsonify({
                'valid': True,
                'user': {
                    'id': admin.id,
                    'username': admin.username,
                    'email': admin.email,
                    'full_name': admin.full_name
                }
            }), 200
        else:
            return jsonify({'valid': False, 'message': 'User not found or inactive'}), 401
    except Exception as e:
        print(f"Token verification error: {e}")
        return jsonify({'valid': False, 'message': 'Invalid token'}), 401



@app.route('/admin/employees', methods=['GET'])
@jwt_required()
def get_employees():
    employees = Employee.query.all()
    return jsonify([{
        'id': emp.id,
        'employee_id': emp.employee_id,
        'name': emp.name,
        'email': emp.email,
        'phone': emp.phone,
        'address': emp.address,
        'department_id': emp.department_id,
        'department_name': emp.department.name if emp.department else None,
        'position': emp.position,
        'hire_date': emp.hire_date.isoformat() if emp.hire_date else None,
        'salary': float(emp.salary) if emp.salary else None,
        'is_active': emp.is_active,
        'created_at': emp.created_at.isoformat(),
        'updated_at': emp.updated_at.isoformat()
    } for emp in employees])

@app.route('/admin/employees', methods=['POST'])
@jwt_required()
def add_employee():
    try:
        data = request.get_json()
        name = data.get('name')
        email = data.get('email')
        employee_id = data.get('employee_id')
        phone = data.get('phone')
        address = data.get('address')
        department_id = data.get('department_id')
        position = data.get('position')
        hire_date_str = data.get('hire_date')
        salary = data.get('salary')
        
        if not name or not email:
            return jsonify({'message': 'Name and email are required'}), 400
        
        # Generate employee_id if not provided
        if not employee_id:
            # Generate EMP001, EMP002, etc.
            last_employee = Employee.query.order_by(Employee.id.desc()).first()
            if last_employee and last_employee.employee_id.startswith('EMP'):
                try:
                    last_num = int(last_employee.employee_id[3:])
                    employee_id = f'EMP{last_num + 1:03d}'
                except:
                    employee_id = 'EMP001'
            else:
                employee_id = 'EMP001'
        
        # Check if employee already exists
        existing_employee = Employee.query.filter(
            (Employee.email == email) | (Employee.employee_id == employee_id)
        ).first()
        if existing_employee:
            return jsonify({'message': 'Employee with this email or employee ID already exists'}), 400
        
        # Parse hire_date
        hire_date = None
        if hire_date_str:
            try:
                hire_date = datetime.strptime(hire_date_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'message': 'Invalid hire_date format. Use YYYY-MM-DD'}), 400
        
        employee = Employee(
            employee_id=employee_id,
            name=name,
            email=email,
            phone=phone,
            address=address,
            department_id=department_id,
            position=position,
            hire_date=hire_date,
            salary=salary
        )
        db.session.add(employee)
        db.session.commit()
        
        # Log the action
        log_audit_action(get_jwt_identity(), 'CREATE', 'employees', employee.id, 
                        None, employee.__dict__, 'Employee created')
        
        return jsonify({
            'id': employee.id,
            'employee_id': employee.employee_id,
            'name': employee.name,
            'email': employee.email,
            'phone': employee.phone,
            'department_id': employee.department_id,
            'position': employee.position,
            'message': 'Employee added successfully'
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"Add employee error: {e}")
        return jsonify({'message': 'Internal server error'}), 500

@app.route('/admin/employees/<int:employee_id>', methods=['PUT'])
@jwt_required()
def update_employee(employee_id):
    try:
        employee = Employee.query.get_or_404(employee_id)
        data = request.get_json()
        
        if not data:
            return jsonify({'message': 'No data provided'}), 400
        
        # Check if email already exists for another employee
        new_email = data.get('email', employee.email)
        if new_email != employee.email:
            existing_employee = Employee.query.filter_by(email=new_email).first()
            if existing_employee:
                return jsonify({'message': 'Employee with this email already exists'}), 400
        
        # Store old values for audit log
        old_values = {
            'name': employee.name,
            'email': employee.email,
            'phone': employee.phone,
            'address': employee.address,
            'department_id': employee.department_id,
            'position': employee.position,
            'hire_date': employee.hire_date.isoformat() if employee.hire_date else None,
            'salary': float(employee.salary) if employee.salary else None
        }
        
        # Update fields
        employee.name = data.get('name', employee.name)
        employee.email = new_email
        employee.phone = data.get('phone', employee.phone)
        employee.address = data.get('address', employee.address)
        employee.department_id = data.get('department_id', employee.department_id)
        employee.position = data.get('position', employee.position)
        employee.salary = data.get('salary', employee.salary)
        
        # Handle hire_date
        hire_date_str = data.get('hire_date')
        if hire_date_str:
            try:
                employee.hire_date = datetime.strptime(hire_date_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'message': 'Invalid hire_date format. Use YYYY-MM-DD'}), 400
        
        db.session.commit()
        
        # Log the action
        new_values = {
            'name': employee.name,
            'email': employee.email,
            'phone': employee.phone,
            'address': employee.address,
            'department_id': employee.department_id,
            'position': employee.position,
            'hire_date': employee.hire_date.isoformat() if employee.hire_date else None,
            'salary': float(employee.salary) if employee.salary else None
        }
        log_audit_action(get_jwt_identity(), 'UPDATE', 'employees', employee.id, 
                        old_values, new_values, 'Employee updated')
        
        return jsonify({
            'id': employee.id,
            'employee_id': employee.employee_id,
            'name': employee.name,
            'email': employee.email,
            'phone': employee.phone,
            'department_id': employee.department_id,
            'department_name': employee.department.name if employee.department else None,
            'position': employee.position,
            'hire_date': employee.hire_date.isoformat() if employee.hire_date else None,
            'salary': float(employee.salary) if employee.salary else None,
            'message': 'Employee updated successfully'
        })
    except Exception as e:
        print(f"Update employee error: {e}")
        return jsonify({'message': 'Internal server error'}), 500

@app.route('/admin/employees/<int:employee_id>', methods=['DELETE'])
@jwt_required()
def delete_employee(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    db.session.delete(employee)
    db.session.commit()
    
    return jsonify({'message': 'Employee deleted successfully'})

@app.route('/admin/attendance', methods=['POST'])
@jwt_required()
def mark_attendance():
    data = request.get_json()
    employee_id = data.get('employee_id')
    status = data.get('status')
    date_str = data.get('date', datetime.now().date().isoformat())
    
    # Convert string date to date object
    if isinstance(date_str, str):
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    else:
        date_obj = date_str
    
    if not employee_id or not status:
        return jsonify({'message': 'Employee ID and status are required'}), 400
    
    valid_statuses = ['present', 'absent', 'half_day', 'leave', 'overtime']
    if status not in valid_statuses:
        return jsonify({'message': f'Status must be one of: {valid_statuses}'}), 400
    
    # Check if attendance already marked for this employee on this date
    existing_attendance = Attendance.query.filter_by(
        employee_id=employee_id, 
        date=date_obj
    ).first()
    
    if existing_attendance:
        existing_attendance.status = status
    else:
        attendance = Attendance(
            employee_id=employee_id,
            date=date_obj,
            status=status
        )
        db.session.add(attendance)
    
    db.session.commit()
    
    return jsonify({'message': 'Attendance marked successfully'})

@app.route('/admin/attendance/report', methods=['GET'])
@jwt_required()
def get_attendance_report():
    date_str = request.args.get('date', datetime.now().date().isoformat())
    
    # Convert string date to date object
    if isinstance(date_str, str):
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    else:
        date_obj = date_str
    
    # Get all employees
    employees = Employee.query.all()
    
    # Get attendance for the specified date
    attendance_records = Attendance.query.filter_by(date=date_obj).all()
    
    # Create a dictionary for quick lookup
    attendance_dict = {record.employee_id: record.status for record in attendance_records}
    
    # Prepare report data
    report_data = []
    present_count = 0
    absent_count = 0
    
    for employee in employees:
        status = attendance_dict.get(employee.id, 'not_marked')
        if status == 'present':
            present_count += 1
        elif status == 'absent':
            absent_count += 1
            
        report_data.append({
            'id': employee.id,
            'name': employee.name,
            'email': employee.email,
            'status': status
        })
    
    return jsonify({
        'date': date_obj.isoformat(),
        'total_employees': len(employees),
        'present_count': present_count,
        'absent_count': absent_count,
        'not_marked_count': len(employees) - present_count - absent_count,
        'employees': report_data
    })

@app.route('/admin/attendance/bulk', methods=['POST'])
@jwt_required()
def bulk_mark_attendance():
    data = request.get_json()
    attendance_data = data.get('attendance_data', [])
    date_str = data.get('date', datetime.now().date().isoformat())
    
    # Convert string date to date object
    if isinstance(date_str, str):
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    else:
        date_obj = date_str
    
    if not attendance_data:
        return jsonify({'message': 'Attendance data is required'}), 400
    
    # Clear existing attendance for this date
    Attendance.query.filter_by(date=date_obj).delete()
    
    # Add new attendance records
    for record in attendance_data:
        attendance = Attendance(
            employee_id=record['employee_id'],
            date=date_obj,
            status=record['status']
        )
        db.session.add(attendance)
    
    db.session.commit()
    
    return jsonify({'message': 'Bulk attendance marked successfully'})

@app.route('/admin/attendance/<int:employee_id>/<string:date>', methods=['DELETE'])
@jwt_required()
def delete_attendance_record(employee_id, date):
    """Delete a specific attendance record"""
    try:
        # Convert string date to date object
        if isinstance(date, str):
            date_obj = datetime.strptime(date, '%Y-%m-%d').date()
        else:
            date_obj = date
        
        # Find and delete the attendance record
        attendance_record = Attendance.query.filter_by(
            employee_id=employee_id,
            date=date_obj
        ).first()
        
        if attendance_record:
            db.session.delete(attendance_record)
            db.session.commit()
            
            # Log the action
            log_audit_action(get_jwt_identity(), 'DELETE', 'attendance', attendance_record.id,
                           {'employee_id': employee_id, 'date': date, 'status': attendance_record.status},
                           None, f'Attendance record deleted for employee {employee_id} on {date}')
            
            return jsonify({'message': 'Attendance record deleted successfully'})
        else:
            return jsonify({'message': 'Attendance record not found'}), 404
            
    except Exception as e:
        db.session.rollback()
        print(f"Delete attendance error: {e}")
        return jsonify({'message': 'Internal server error'}), 500

@app.route('/admin/attendance/date/<string:date>', methods=['DELETE'])
@jwt_required()
def clear_attendance_by_date(date):
    """Clear all attendance records for a specific date"""
    try:
        # Convert string date to date object
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
        
        # Count records before deletion for logging
        records_to_delete = Attendance.query.filter_by(date=date_obj).all()
        deleted_count = len(records_to_delete)
        
        # Delete all attendance records for the date
        Attendance.query.filter_by(date=date_obj).delete()
        
        db.session.commit()
        
        # Log the action
        log_audit_action(get_jwt_identity(), 'DELETE', 'attendance', None,
                       None, {'date': date, 'deleted_count': deleted_count},
                       f'Cleared {deleted_count} attendance records for {date}')
        
        return jsonify({
            'message': f'Successfully cleared {deleted_count} attendance records for {date}',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Clear attendance by date error: {e}")
        return jsonify({'message': 'Internal server error'}), 500

@app.route('/admin/attendance/month/<string:date>', methods=['DELETE'])
@jwt_required()
def clear_monthly_attendance(date):
    """Clear all attendance records for a specific month"""
    try:
        # Convert string date to date object and get month boundaries
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
        first_day = date_obj.replace(day=1)
        if first_day.month == 12:
            last_day = first_day.replace(year=first_day.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last_day = first_day.replace(month=first_day.month + 1, day=1) - timedelta(days=1)
        
        # Count records before deletion for logging
        records_to_delete = Attendance.query.filter(
            Attendance.date >= first_day,
            Attendance.date <= last_day
        ).all()
        
        deleted_count = len(records_to_delete)
        
        # Delete all attendance records for the month
        Attendance.query.filter(
            Attendance.date >= first_day,
            Attendance.date <= last_day
        ).delete()
        
        db.session.commit()
        
        # Log the action
        log_audit_action(get_jwt_identity(), 'DELETE', 'attendance', None,
                       None, {'month': first_day.strftime('%Y-%m'), 'deleted_count': deleted_count},
                       f'Cleared {deleted_count} attendance records for {first_day.strftime("%B %Y")}')
        
        return jsonify({
            'message': f'Successfully cleared {deleted_count} attendance records for {first_day.strftime("%B %Y")}',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Clear monthly attendance error: {e}")
        return jsonify({'message': 'Internal server error'}), 500

@app.route('/admin/attendance/overview', methods=['GET'])
@jwt_required()
def get_attendance_overview():
    """Get monthly attendance overview for all employees"""
    try:
        date_str = request.args.get('date', datetime.now().date().isoformat())
        
        # Convert to date object and get month boundaries
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        first_day = date_obj.replace(day=1)
        if first_day.month == 12:
            last_day = first_day.replace(year=first_day.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last_day = first_day.replace(month=first_day.month + 1, day=1) - timedelta(days=1)
        
        # Get all active employees
        employees = Employee.query.filter_by(is_active=True).all()
        
        # Get all attendance records for the month
        attendance_records = Attendance.query.filter(
            Attendance.date >= first_day,
            Attendance.date <= last_day
        ).all()
        
        # Create attendance lookup dictionary
        attendance_dict = {}
        for record in attendance_records:
            employee_id = record.employee_id
            date_key = record.date.isoformat()
            
            if employee_id not in attendance_dict:
                attendance_dict[employee_id] = {}
            
            attendance_dict[employee_id][date_key] = record.status
        
        return jsonify({
            'month': first_day.strftime('%Y-%m'),
            'first_day': first_day.isoformat(),
            'last_day': last_day.isoformat(),
            'employees': [{
                'id': emp.id,
                'employee_id': emp.employee_id,
                'name': emp.name,
                'email': emp.email,
                'department': emp.department.name if emp.department else None
            } for emp in employees],
            'attendance_data': attendance_dict
        })
        
    except Exception as e:
        print(f"Attendance overview error: {e}")
        return jsonify({'error': 'Failed to fetch attendance overview'}), 500

@app.route('/admin/test-token', methods=['GET'])
@jwt_required()
def test_token():
    """Test endpoint to verify JWT token"""
    current_user_id = get_jwt_identity()
    return jsonify({
        'message': 'Token is valid',
        'user_id': current_user_id
    })

@app.route('/admin/attendance/validate', methods=['GET'])
@jwt_required()
def validate_attendance_completion():
    """Validate if all attendance is marked for the specified month up to today"""
    try:
        date_str = request.args.get('date', datetime.now().date().isoformat())
        
        # Convert string date to date object
        if isinstance(date_str, str):
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            date_obj = date_str
            
        # Get first day of the month
        first_day = date_obj.replace(day=1)
        today = datetime.now().date()
        
        # Only validate up to today if we're in the current month
        if first_day.year == today.year and first_day.month == today.month:
            last_day = today
        else:
            # For past/future months, get the last day of that month
            if first_day.month == 12:
                last_day = first_day.replace(year=first_day.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                last_day = first_day.replace(month=first_day.month + 1, day=1) - timedelta(days=1)
        
        # Get all active employees
        employees = Employee.query.filter_by(is_active=True).all()
        
        # Get holidays for the period
        holidays = Holiday.query.filter(
            Holiday.date >= first_day,
            Holiday.date <= last_day
        ).all()
        holiday_dates = {holiday.date for holiday in holidays}
        
        # Get attendance records for the period
        attendance_records = Attendance.query.filter(
            Attendance.date >= first_day,
            Attendance.date <= last_day
        ).all()
        
        # Create attendance lookup
        attendance_dict = {}
        for record in attendance_records:
            key = f"{record.employee_id}_{record.date.isoformat()}"
            attendance_dict[key] = record.status
        
        # Count working days (excluding weekends and holidays)
        working_days = []
        current_date = first_day
        while current_date <= last_day:
            if current_date.weekday() < 5 and current_date not in holiday_dates:  # Exclude weekends and holidays
                working_days.append(current_date)
            current_date += timedelta(days=1)
        
        # Check for missing attendance
        missing_attendance = []
        total_expected_records = len(employees) * len(working_days)
        actual_records = 0
        
        for employee in employees:
            for work_day in working_days:
                key = f"{employee.id}_{work_day.isoformat()}"
                if key in attendance_dict:
                    actual_records += 1
                else:
                    missing_attendance.append({
                        'employee_id': employee.id,
                        'employee_name': employee.name,
                        'date': work_day.isoformat(),
                        'date_formatted': work_day.strftime('%B %d, %Y')
                    })
        
        completion_percentage = (actual_records / total_expected_records * 100) if total_expected_records > 0 else 100
        
        return jsonify({
            'is_complete': len(missing_attendance) == 0,
            'total_expected': total_expected_records,
            'total_marked': actual_records,
            'missing_count': len(missing_attendance),
            'completion_percentage': round(completion_percentage, 1),
            'missing_attendance': missing_attendance[:10],  # Limit to first 10 for display
            'period': {
                'start_date': first_day.isoformat(),
                'end_date': last_day.isoformat(),
                'month_name': first_day.strftime('%B %Y')
            },
            'working_days_count': len(working_days),
            'employees_count': len(employees)
        })
        
    except Exception as e:
        print(f"Attendance validation error: {e}")
        return jsonify({'error': f'Failed to validate attendance: {str(e)}'}), 500

@app.route('/admin/attendance/export', methods=['GET'])
@jwt_required()
def export_attendance_monthly_report():
    """Export full month attendance report to Excel file"""
    try:
        date_str = request.args.get('date', datetime.now().date().isoformat())
        print(f"Monthly export request for date: {date_str}")
        
        # Convert string date to date object and get month start/end
        if isinstance(date_str, str):
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            date_obj = date_str
            
        # Get first day of the month and last day
        first_day = date_obj.replace(day=1)
        today = datetime.now().date()
        
        # Check if user wants to force full month export
        force_full_month = request.args.get('force_full_month', 'false').lower() == 'true'
        
        # Calculate the last day of the month
        if first_day.month == 12:
            month_last_day = first_day.replace(year=first_day.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            month_last_day = first_day.replace(month=first_day.month + 1, day=1) - timedelta(days=1)
        
        # Determine export range
        if force_full_month:
            last_day = month_last_day  # Force full month when requested
            print(f"Force full month export: {first_day} to {last_day}")
        elif first_day.year == today.year and first_day.month == today.month:
            last_day = today  # Only show dates up to today for current month
            print(f"Current month partial export: {first_day} to {last_day}")
        else:
            last_day = month_last_day  # Show full month for past/future months
            print(f"Full month export: {first_day} to {last_day}")
        
        # Get all employees
        employees = Employee.query.filter_by(is_active=True).all()
        print(f"Found {len(employees)} active employees")
        
        # Get all attendance records for the month
        attendance_records = Attendance.query.filter(
            Attendance.date >= first_day,
            Attendance.date <= last_day
        ).all()
        print(f"Found {len(attendance_records)} attendance records for the month")
        
        # Get holidays for the month
        try:
            holidays = Holiday.query.filter(
                Holiday.date >= first_day,
                Holiday.date <= last_day
            ).all()
            print(f"Found {len(holidays)} holidays for the month")
        except Exception as e:
            print(f"Error fetching holidays: {e}")
            holidays = []
        
        # Create holiday lookup dictionary
        holiday_dict = {}
        for holiday in holidays:
            holiday_dict[holiday.date.isoformat()] = holiday.name
            print(f"Holiday mapped: {holiday.date.isoformat()} -> {holiday.name}")
        
        # Create attendance lookup dictionary
        attendance_dict = {}
        for record in attendance_records:
            key = f"{record.employee_id}_{record.date.isoformat()}"
            attendance_dict[key] = record.status
        
        # Generate all dates in the month
        current_date = first_day
        all_dates = []
        while current_date <= last_day:
            all_dates.append(current_date)
            current_date += timedelta(days=1)
        
        # Create Excel workbook with compatibility settings
        wb = Workbook()
        ws = wb.active
        # Ensure worksheet title is Excel-safe
        safe_title = f"Attendance {first_day.strftime('%B %Y')}"
        # Remove any characters that might cause issues
        safe_title = ''.join(c for c in safe_title if c.isalnum() or c in ' -_')
        ws.title = safe_title[:31]  # Excel sheet names max 31 chars
        
        # Set workbook properties for compatibility
        wb.properties.title = f"Attendance Overview - {first_day.strftime('%B %Y')}"
        wb.properties.subject = "Employee Attendance Report"
        wb.properties.creator = "Attendance Management System"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        weekend_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # Status color mapping
        status_fills = {
            'present': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
            'half_day': PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),
            'absent': PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),
            'leave': PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
            'overtime': PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
        }
        
        # Holiday fill pattern (red background)
        holiday_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # Create headers
        headers = ['Employee ID', 'Employee Name', 'Email', 'Department']
        for date in all_dates:
            # Remove newlines to avoid Excel warnings
            headers.append(f"{date.day} {date.strftime('%a')}")
        headers.extend(['Present', 'Half Day', 'Absent', 'Leave', 'Overtime', 'Total Working Days'])
        
        # Set headers with improved formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=str(header))  # Ensure string value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Set header row height
        ws.row_dimensions[1].height = 25
        
        # Add employee data
        for row, employee in enumerate(employees, 2):
            # Employee info - ensure all values are clean strings
            employee_id = str(employee.id) if employee.id else 'N/A'
            employee_name = str(employee.name).strip() if employee.name else 'N/A'
            employee_email = str(employee.email).strip() if employee.email else 'N/A'
            
            # Handle department - check if it's a relationship object or string
            if hasattr(employee, 'department') and employee.department:
                if hasattr(employee.department, 'name'):
                    dept_name = str(employee.department.name).strip() if employee.department.name else 'N/A'
                else:
                    dept_name = str(employee.department).strip()
            else:
                dept_name = 'N/A'
            
            ws.cell(row=row, column=1, value=employee_id)
            ws.cell(row=row, column=2, value=employee_name)
            ws.cell(row=row, column=3, value=employee_email)
            ws.cell(row=row, column=4, value=dept_name)
            
            # Attendance data for each day
            stats = {'present': 0, 'half_day': 0, 'absent': 0, 'leave': 0, 'overtime': 0}
            
            for col_idx, date in enumerate(all_dates, 5):
                key = f"{employee.id}_{date.isoformat()}"
                status = attendance_dict.get(key, '')
                date_str = date.isoformat()
                is_holiday = date_str in holiday_dict
                holiday_name = holiday_dict.get(date_str, '')
                
                # Determine display value and fill
                if is_holiday:
                    # Show holiday name (keep it concise for Excel cells)
                    if len(holiday_name) > 15:
                        display_value = f"Holiday: {holiday_name[:12]}..."
                    else:
                        display_value = f"Holiday: {holiday_name}"
                    cell_fill = holiday_fill
                elif status:
                    display_value = status.replace('_', ' ').title()
                    cell_fill = status_fills.get(status, None)
                    if status in stats:
                        stats[status] += 1
                elif date.weekday() >= 5:  # Weekend
                    display_value = ''
                    cell_fill = weekend_fill
                else:
                    display_value = ''
                    cell_fill = None
                
                cell = ws.cell(row=row, column=col_idx, value=display_value)
                cell.alignment = center_alignment
                
                # Apply the determined fill
                if cell_fill:
                    cell.fill = cell_fill
            
            # Add summary statistics with proper numeric formatting
            stats_start_col = len(all_dates) + 5
            ws.cell(row=row, column=stats_start_col, value=int(stats['present']))
            ws.cell(row=row, column=stats_start_col + 1, value=int(stats['half_day']))
            ws.cell(row=row, column=stats_start_col + 2, value=int(stats['absent']))
            ws.cell(row=row, column=stats_start_col + 3, value=int(stats['leave']))
            ws.cell(row=row, column=stats_start_col + 4, value=int(stats['overtime']))
            ws.cell(row=row, column=stats_start_col + 5, value=int(sum(stats.values())))
        
        # Auto-adjust column widths safely
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    continue  # Skip problematic cells
            # Ensure reasonable width bounds
            adjusted_width = min(max(max_length + 2, 10), 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO buffer with proper handling
        output = io.BytesIO()
        try:
            wb.save(output)
            output.seek(0)
            file_data = output.getvalue()
            output.seek(0)  # Reset for send_file
        except Exception as save_error:
            print(f"Error saving workbook: {save_error}")
            raise
        
        # Create filename
        filename = f"attendance_overview_{first_day.strftime('%Y%m')}.xlsx"
        print(f"Excel file created: {filename}")
        
        # Save file to database
        try:
            file_id = save_file_to_db(
                file_data=file_data,
                filename=filename,
                file_type='excel',
                description=f'Monthly attendance overview for {first_day.strftime("%B %Y")}',
                related_table='attendance',
                related_id=None
            )
            
            # Log the export action
            log_audit_action(get_jwt_identity(), 'EXPORT', 'attendance', None, 
                           None, {'filename': filename, 'file_id': file_id}, 
                           f'Monthly Excel report exported for {first_day.strftime("%B %Y")}')
            
            print(f"File saved to database with ID: {file_id}")
        except Exception as e:
            print(f"Error saving file to database: {e}")
    
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"Excel export error: {e}")
        return jsonify({'error': f'Failed to export Excel: {str(e)}'}), 500

@app.route('/admin/attendance/export-pdf', methods=['GET'])
@jwt_required()
def export_attendance_monthly_report_pdf():
    """Export monthly attendance report to PDF file"""
    try:
        date_str = request.args.get('date', datetime.now().date().isoformat())
        print(f"Monthly PDF Export request for date: {date_str}")
        
        # Convert string date to date object and get month start/end
        if isinstance(date_str, str):
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            date_obj = date_str
            
        # Get first and last day of the month
        first_day = date_obj.replace(day=1)
        if first_day.month == 12:
            last_day = first_day.replace(year=first_day.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last_day = first_day.replace(month=first_day.month + 1, day=1) - timedelta(days=1)
        
        # Get all active employees
        employees = Employee.query.filter_by(is_active=True).all()
        
        # Get all attendance records for the month
        attendance_records = Attendance.query.filter(
            Attendance.date >= first_day,
            Attendance.date <= last_day
        ).all()
        
        # Create attendance lookup dictionary
        attendance_dict = {}
        for record in attendance_records:
            key = f"{record.employee_id}_{record.date.isoformat()}"
            attendance_dict[key] = record.status
        
        # Generate all dates in the month
        current_date = first_day
        all_dates = []
        while current_date <= last_day:
            all_dates.append(current_date)
            current_date += timedelta(days=1)
        
        # Create PDF in memory
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=1  # Center alignment
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=8,
            leading=10
        )
        
        # Create content
        story = []
        
        # Title
        title = f"Monthly Attendance Overview - {first_day.strftime('%B %Y')}"
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 20))
        
        # Summary statistics
        total_employees = len(employees)
        total_working_days = len([d for d in all_dates if d.weekday() < 5])  # Exclude weekends
        
        summary_text = f"<b>Report Summary:</b><br/>"
        summary_text += f"Total Employees: {total_employees}<br/>"
        summary_text += f"Total Working Days: {total_working_days}<br/>"
        summary_text += f"Report Period: {first_day.strftime('%B %d')} - {last_day.strftime('%B %d, %Y')}<br/>"
        
        story.append(Paragraph(summary_text, normal_style))
        story.append(Spacer(1, 20))
        
        # Create summary table for each employee
        table_data = [['Employee', 'Present', 'Half Day', 'Absent', 'Leave', 'Overtime', 'Total Days', 'Attendance %']]
        
        for employee in employees:
            stats = {'present': 0, 'half_day': 0, 'absent': 0, 'leave': 0, 'overtime': 0}
            
            for date in all_dates:
                if date.weekday() < 5:  # Only count working days
                    key = f"{employee.id}_{date.isoformat()}"
                    status = attendance_dict.get(key, '')
                    if status in stats:
                        stats[status] += 1
            
            total_marked = sum(stats.values())
            attendance_percentage = (stats['present'] / total_working_days * 100) if total_working_days > 0 else 0
            
            table_data.append([
                employee.name,
                str(stats['present']),
                str(stats['half_day']),
                str(stats['absent']),
                str(stats['leave']),
                str(stats['overtime']),
                str(total_marked),
                f"{attendance_percentage:.1f}%"
            ])
        
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
        
        # Add legend
        legend_text = "<b>Status Legend:</b><br/>"
        legend_text += "Present: Full working day<br/>"
        legend_text += "Half Day: Partial working day<br/>"
        legend_text += "Absent: Did not attend<br/>"
        legend_text += "Leave: On approved leave<br/>"
        legend_text += "Overtime: Worked extra hours<br/>"
        
        story.append(Paragraph(legend_text, normal_style))
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        # Create filename
        filename = f"attendance_overview_{first_day.strftime('%Y%m')}.pdf"
        
        # Save file to database (optional)
        try:
            file_id = save_file_to_db(
                file_data=output.getvalue(),
                filename=filename,
                file_type='pdf',
                description=f'Monthly attendance overview PDF for {first_day.strftime("%B %Y")}',
                related_table='attendance',
                related_id=None
            )
            
            # Log the export action
            log_audit_action(get_jwt_identity(), 'EXPORT', 'attendance', None, 
                           None, {'filename': filename, 'file_id': file_id}, 
                           f'Monthly PDF report exported for {first_day.strftime("%B %Y")}')
        except Exception as e:
            print(f"Error saving PDF to database: {e}")
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
    
    except Exception as e:
        print(f"PDF Export error: {str(e)}")
        return jsonify({'error': 'Failed to export PDF report'}), 500

# Department Management Endpoints
@app.route('/admin/departments', methods=['GET'])
@jwt_required()
def get_departments():
    include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'
    
    if include_inactive:
        departments = Department.query.all()
    else:
        departments = Department.query.filter_by(is_active=True).all()
    
    return jsonify([{
        'id': dept.id,
        'name': dept.name,
        'description': dept.description,
        'manager_id': dept.manager_id,
        'manager_name': dept.manager.name if dept.manager else None,
        'is_active': dept.is_active,
        'employee_count': len([emp for emp in dept.employees if emp.is_active]),
        'created_at': dept.created_at.isoformat(),
        'updated_at': dept.updated_at.isoformat()
    } for dept in departments])

@app.route('/admin/departments', methods=['POST'])
@jwt_required()
def add_department():
    try:
        data = request.get_json()
        name = data.get('name')
        description = data.get('description')
        manager_id = data.get('manager_id')
        is_active = data.get('is_active', True)
        
        if not name:
            return jsonify({'message': 'Department name is required'}), 400
        
        # Check if department already exists
        existing_dept = Department.query.filter_by(name=name).first()
        if existing_dept:
            return jsonify({'message': 'Department with this name already exists'}), 400
        
        # Validate manager if provided
        if manager_id:
            manager = Employee.query.filter_by(id=manager_id, is_active=True).first()
            if not manager:
                return jsonify({'message': 'Invalid manager selected'}), 400
        
        department = Department(
            name=name,
            description=description,
            manager_id=manager_id,
            is_active=is_active
        )
        db.session.add(department)
        db.session.commit()
        
        # Log the action
        log_audit_action(get_jwt_identity(), 'CREATE', 'departments', department.id, 
                        None, {
                            'name': name, 
                            'description': description, 
                            'manager_id': manager_id,
                            'is_active': is_active
                        }, 'Department created')
        
        return jsonify({
            'id': department.id,
            'name': department.name,
            'description': department.description,
            'manager_id': department.manager_id,
            'manager_name': department.manager.name if department.manager else None,
            'is_active': department.is_active,
            'message': 'Department added successfully'
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"Add department error: {e}")
        return jsonify({'message': 'Internal server error'}), 500

@app.route('/admin/departments/<int:department_id>', methods=['PUT'])
@jwt_required()
def update_department(department_id):
    try:
        department = Department.query.get_or_404(department_id)
        data = request.get_json()
        
        if not data:
            return jsonify({'message': 'No data provided'}), 400
        
        name = data.get('name')
        description = data.get('description')
        manager_id = data.get('manager_id')
        is_active = data.get('is_active')
        
        if not name:
            return jsonify({'message': 'Department name is required'}), 400
        
        # Check if another department with the same name exists
        if name != department.name:
            existing_dept = Department.query.filter_by(name=name).first()
            if existing_dept:
                return jsonify({'message': 'Department with this name already exists'}), 400
        
        # Validate manager if provided
        if manager_id:
            manager = Employee.query.filter_by(id=manager_id, is_active=True).first()
            if not manager:
                return jsonify({'message': 'Invalid manager selected'}), 400
        
        old_values = {
            'name': department.name,
            'description': department.description,
            'manager_id': department.manager_id,
            'is_active': department.is_active
        }
        
        department.name = name
        department.description = description
        if manager_id is not None:  # Allow setting to None
            department.manager_id = manager_id
        if is_active is not None:
            department.is_active = is_active
        
        db.session.commit()
        
        # Log the action
        new_values = {
            'name': name, 
            'description': description,
            'manager_id': department.manager_id,
            'is_active': department.is_active
        }
        log_audit_action(get_jwt_identity(), 'UPDATE', 'departments', department.id, 
                        old_values, new_values, 'Department updated')
        
        return jsonify({
            'id': department.id,
            'name': department.name,
            'description': department.description,
            'manager_id': department.manager_id,
            'manager_name': department.manager.name if department.manager else None,
            'is_active': department.is_active,
            'message': 'Department updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Update department error: {e}")
        return jsonify({'message': 'Internal server error'}), 500

@app.route('/admin/departments/<int:department_id>', methods=['DELETE'])
@jwt_required()
def delete_department(department_id):
    try:
        department = Department.query.get_or_404(department_id)
        
        # Check if department has employees
        if department.employees:
            return jsonify({
                'message': f'Cannot delete department "{department.name}" because it has {len(department.employees)} employees assigned to it. Please reassign employees first.'
            }), 400
        
        department_name = department.name
        db.session.delete(department)
        db.session.commit()
        
        # Log the action
        log_audit_action(get_jwt_identity(), 'DELETE', 'departments', department_id, 
                        {'name': department_name}, None, f'Department "{department_name}" deleted')
        
        return jsonify({'message': f'Department "{department_name}" deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Delete department error: {e}")
        return jsonify({'message': 'Internal server error'}), 500

@app.route('/admin/employees/for-manager', methods=['GET'])
@jwt_required()
def get_employees_for_manager():
    """Get active employees who can be assigned as department managers"""
    try:
        employees = Employee.query.filter_by(is_active=True).all()
        return jsonify([{
            'id': emp.id,
            'name': emp.name,
            'employee_id': emp.employee_id,
            'position': emp.position,
            'department_name': emp.department.name if emp.department else 'No Department'
        } for emp in employees])
    except Exception as e:
        print(f"Get employees for manager error: {e}")
        return jsonify({'message': 'Internal server error'}), 500

# File Management Endpoints
@app.route('/admin/files', methods=['GET'])
@jwt_required()
def get_files():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    file_type = request.args.get('type')
    
    query = FileStorage.query
    if file_type:
        query = query.filter_by(file_type=file_type)
    
    files = query.order_by(FileStorage.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'files': [{
            'id': f.id,
            'filename': f.original_filename,
            'file_type': f.file_type,
            'file_size': f.file_size,
            'description': f.description,
            'uploaded_by': f.uploader.username if f.uploader else None,
            'created_at': f.created_at.isoformat()
        } for f in files.items],
        'total': files.total,
        'pages': files.pages,
        'current_page': page
    })

@app.route('/admin/files/<int:file_id>/download', methods=['GET'])
@jwt_required()
def download_file(file_id):
    try:
        file_record = FileStorage.query.get_or_404(file_id)
        
        # Log the download
        log_audit_action(get_jwt_identity(), 'DOWNLOAD', 'file_storage', file_id, 
                        None, {'filename': file_record.original_filename}, 'File downloaded')
        
        return send_file(
            io.BytesIO(file_record.file_data),
            as_attachment=True,
            download_name=file_record.original_filename,
            mimetype=file_record.mime_type
        )
        
    except Exception as e:
        print(f"File download error: {e}")
        return jsonify({'message': 'File not found or error occurred'}), 404

# Leave Management Endpoints
@app.route('/admin/leaves', methods=['GET'])
@jwt_required()
def get_leaves():
    leaves = Leave.query.order_by(Leave.created_at.desc()).all()
    return jsonify([{
        'id': leave.id,
        'employee_id': leave.employee_id,
        'employee_name': leave.employee.name,
        'leave_type': leave.leave_type,
        'start_date': leave.start_date.isoformat(),
        'end_date': leave.end_date.isoformat(),
        'days_count': leave.days_count,
        'reason': leave.reason,
        'status': leave.status,
        'created_at': leave.created_at.isoformat()
    } for leave in leaves])

@app.route('/admin/leaves/<int:leave_id>/approve', methods=['PUT'])
@jwt_required()
def approve_leave(leave_id):
    try:
        leave = Leave.query.get_or_404(leave_id)
        action = request.get_json().get('action')  # 'approve' or 'reject'
        
        if action not in ['approve', 'reject']:
            return jsonify({'message': 'Invalid action'}), 400
        
        old_status = leave.status
        leave.status = 'approved' if action == 'approve' else 'rejected'
        leave.approved_by = get_jwt_identity()
        leave.approved_at = datetime.utcnow()
        
        db.session.commit()
        
        # Log the action
        log_audit_action(get_jwt_identity(), 'UPDATE', 'leaves', leave_id, 
                        {'status': old_status}, {'status': leave.status}, 
                        f'Leave {action}d')
        
        return jsonify({
            'message': f'Leave {action}d successfully',
            'leave': {
                'id': leave.id,
                'status': leave.status,
                'approved_at': leave.approved_at.isoformat() if leave.approved_at else None
            }
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Leave approval error: {e}")
        return jsonify({'message': 'Internal server error'}), 500

# Holiday Management Endpoints
@app.route('/admin/holidays', methods=['GET'])
@jwt_required()
def get_holidays():
    """Get all holidays"""
    try:
        holidays = Holiday.query.order_by(Holiday.date.asc()).all()
        return jsonify([{
            'id': holiday.id,
            'name': holiday.name,
            'date': holiday.date.isoformat(),
            'description': holiday.description,
            'is_recurring': holiday.is_recurring,
            'created_by': holiday.creator.username if holiday.creator else None,
            'created_at': holiday.created_at.isoformat()
        } for holiday in holidays])
    except Exception as e:
        print(f"Get holidays error: {e}")
        return jsonify({'message': 'Internal server error'}), 500

@app.route('/admin/holidays', methods=['POST'])
@jwt_required()
def add_holiday():
    """Add a new holiday"""
    try:
        data = request.get_json()
        name = data.get('name')
        date_str = data.get('date')
        description = data.get('description', '')
        is_recurring = data.get('is_recurring', False)
        
        if not name or not date_str:
            return jsonify({'message': 'Holiday name and date are required'}), 400
        
        # Parse the date
        try:
            holiday_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'message': 'Invalid date format. Use YYYY-MM-DD'}), 400
        
        # Check if holiday already exists on this date
        existing_holiday = Holiday.query.filter_by(date=holiday_date).first()
        if existing_holiday:
            return jsonify({
                'message': f'A holiday "{existing_holiday.name}" already exists on {date_str}'
            }), 400
        
        # Create new holiday
        holiday = Holiday(
            name=name.strip(),
            date=holiday_date,
            description=description.strip() if description else None,
            is_recurring=is_recurring,
            created_by=get_jwt_identity()
        )
        
        db.session.add(holiday)
        db.session.commit()
        
        # Log the action
        log_audit_action(get_jwt_identity(), 'CREATE', 'holidays', holiday.id, 
                        None, {
                            'name': name,
                            'date': date_str,
                            'description': description,
                            'is_recurring': is_recurring
                        }, f'Holiday "{name}" created')
        
        return jsonify({
            'id': holiday.id,
            'name': holiday.name,
            'date': holiday.date.isoformat(),
            'description': holiday.description,
            'is_recurring': holiday.is_recurring,
            'created_by': holiday.creator.username if holiday.creator else None,
            'created_at': holiday.created_at.isoformat(),
            'message': f'Holiday "{name}" added successfully'
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"Add holiday error: {e}")
        return jsonify({'message': 'Internal server error'}), 500

@app.route('/admin/holidays/<int:holiday_id>', methods=['PUT'])
@jwt_required()
def update_holiday(holiday_id):
    """Update an existing holiday"""
    try:
        holiday = Holiday.query.get_or_404(holiday_id)
        data = request.get_json()
        
        if not data:
            return jsonify({'message': 'No data provided'}), 400
        
        name = data.get('name', holiday.name)
        date_str = data.get('date')
        description = data.get('description', holiday.description)
        is_recurring = data.get('is_recurring', holiday.is_recurring)
        
        # Store old values for audit log
        old_values = {
            'name': holiday.name,
            'date': holiday.date.isoformat(),
            'description': holiday.description,
            'is_recurring': holiday.is_recurring
        }
        
        # Parse new date if provided
        if date_str:
            try:
                holiday_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'message': 'Invalid date format. Use YYYY-MM-DD'}), 400
            
            # Check if another holiday exists on this date (excluding current holiday)
            if holiday_date != holiday.date:
                existing_holiday = Holiday.query.filter(
                    Holiday.date == holiday_date,
                    Holiday.id != holiday_id
                ).first()
                if existing_holiday:
                    return jsonify({
                        'message': f'A holiday "{existing_holiday.name}" already exists on {date_str}'
                    }), 400
        else:
            holiday_date = holiday.date
        
        # Update holiday
        holiday.name = name.strip()
        holiday.date = holiday_date
        holiday.description = description.strip() if description else None
        holiday.is_recurring = is_recurring
        
        db.session.commit()
        
        # Log the action
        new_values = {
            'name': holiday.name,
            'date': holiday.date.isoformat(),
            'description': holiday.description,
            'is_recurring': holiday.is_recurring
        }
        log_audit_action(get_jwt_identity(), 'UPDATE', 'holidays', holiday.id, 
                        old_values, new_values, f'Holiday "{holiday.name}" updated')
        
        return jsonify({
            'id': holiday.id,
            'name': holiday.name,
            'date': holiday.date.isoformat(),
            'description': holiday.description,
            'is_recurring': holiday.is_recurring,
            'created_by': holiday.creator.username if holiday.creator else None,
            'created_at': holiday.created_at.isoformat(),
            'message': f'Holiday "{holiday.name}" updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Update holiday error: {e}")
        return jsonify({'message': 'Internal server error'}), 500

@app.route('/admin/holidays/<int:holiday_id>', methods=['DELETE'])
@jwt_required()
def delete_holiday(holiday_id):
    """Delete a holiday"""
    try:
        holiday = Holiday.query.get_or_404(holiday_id)
        holiday_name = holiday.name
        holiday_date = holiday.date.isoformat()
        
        db.session.delete(holiday)
        db.session.commit()
        
        # Log the action
        log_audit_action(get_jwt_identity(), 'DELETE', 'holidays', holiday_id, 
                        {
                            'name': holiday_name,
                            'date': holiday_date
                        }, None, f'Holiday "{holiday_name}" deleted')
        
        return jsonify({
            'message': f'Holiday "{holiday_name}" deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Delete holiday error: {e}")
        return jsonify({'message': 'Internal server error'}), 500

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # Create default admin if not exists
        admin = Admin.query.filter_by(username='admin').first()
        if not admin:
            admin = Admin(
                username='admin',
                password_hash=generate_password_hash('admin123'),
                email='admin@company.com',
                full_name='System Administrator'
            )
            db.session.add(admin)
            db.session.commit()
            print("Default admin created - Username: admin, Password: admin123")
        
        # Create default department if not exists
        dept = Department.query.filter_by(name='General').first()
        if not dept:
            dept = Department(
                name='General',
                description='General department for unassigned employees'
            )
            db.session.add(dept)
            db.session.commit()
            print("Default department 'General' created")
        
        # Create default holidays if not exists
        current_year = datetime.now().year
        default_holidays = [
            {'name': 'New Year Day', 'date': f'{current_year}-01-01'},
            {'name': 'Christmas Day', 'date': f'{current_year}-12-25'},
        ]
        
        for holiday_data in default_holidays:
            existing_holiday = Holiday.query.filter_by(
                name=holiday_data['name'],
                date=datetime.strptime(holiday_data['date'], '%Y-%m-%d').date()
            ).first()
            
            if not existing_holiday:
                holiday = Holiday(
                    name=holiday_data['name'],
                    date=datetime.strptime(holiday_data['date'], '%Y-%m-%d').date(),
                    description=f"Default {holiday_data['name']} holiday",
                    created_by=admin.id
                )
                db.session.add(holiday)
        
        db.session.commit()
        print(f"Default holidays checked/created for year {current_year}")
    
    app.run(debug=True, host='0.0.0.0', port=5000)
